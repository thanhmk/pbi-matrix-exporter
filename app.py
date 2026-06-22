# -*- coding: utf-8 -*-
"""
PBI Matrix Exporter — paste query tu Performance Analyzer, xuat Excel dung dinh dang.
Ho tro: phan cap hang (1..n cap), phan cap cot (vd Date), nhieu measure, che do dao,
va fallback xuat phang (khong bao gio crash).
Chay: streamlit run app.py   |   pip install streamlit pywin32 pandas openpyxl
"""

import io
import os
import re
import glob
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="PBI Matrix Exporter", page_icon="📊", layout="wide")

THIN = Side(style="thin", color="DBDBDB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR = PatternFill("solid", fgColor="2F5597")
LEVEL_FILL = ["C9DAF1", "E2EAF6", "EFF3FA", "F6F8FB", "FAFBFD"]

# ============================================================
# 1. TIM PORT POWER BI DESKTOP
# ============================================================
def get_pbi_local_port():
    import win32com.client
    try:
        wmi = win32com.client.GetObject("winmgmts:")
        for p in wmi.ExecQuery("Select CommandLine from Win32_Process where Name='msmdsrv.exe'"):
            if p.CommandLine:
                m = re.search(r'-s\s+"([^"]+)"', p.CommandLine, re.IGNORECASE)
                if m:
                    pf = os.path.join(m.group(1), 'msmdsrv.port.txt')
                    if os.path.exists(pf):
                        with open(pf, 'r', encoding='utf-16le') as f:
                            port = f.read().strip()
                            if port:
                                return port
    except Exception:
        pass
    lad = os.environ.get('LOCALAPPDATA', ''); tmp = os.environ.get('TEMP', '')
    paths = [
        os.path.join(lad, 'Microsoft', 'Power BI Desktop', 'AnalysisServicesWorkspaces'),
        os.path.join(lad, 'Packages', 'Microsoft.MicrosoftPowerBIDesktop_8wekyb3d8bbwe',
                     'LocalCache', 'Local', 'Microsoft', 'Power BI Desktop', 'AnalysisServicesWorkspaces'),
        os.path.join(tmp, 'Power BI Desktop', 'AnalysisServicesWorkspaces'),
    ]
    files = []
    for p in paths:
        if os.path.exists(p):
            files.extend(glob.glob(os.path.join(p, '*', 'Data', 'msmdsrv.port.txt')))
    if files:
        with open(max(files, key=os.path.getctime), 'r', encoding='utf-16le') as f:
            return f.read().strip()
    raise RuntimeError("Không tìm thấy Power BI Desktop. Hãy mở sẵn 1 file .pbix.")

# ============================================================
# 2. CHAY QUERY (Hack limit)
# ============================================================
def prep_query(dax):
    dax = re.sub(r'(?i)\bTOPN\s*\(\s*\d+', 'TOPN(1048576', dax)
    if "__DS0Core" in dax:
        head = re.split(r'\bEVALUATE\b', dax, flags=re.IGNORECASE)[0].rstrip()
        return head + "\nEVALUATE\n__DS0Core"
    return dax

def _grab(rs):
    cols = [f.Name for f in rs.Fields]
    rows = []
    if not rs.EOF:
        for row in zip(*rs.GetRows()):
            rows.append([v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else v for v in row])
    return cols, rows

def _has_dim(cols):
    return any(('[' in c and not c.lstrip().startswith('[')) for c in cols)

def run_query(dax):
    import win32com.client
    port = get_pbi_local_port()
    conn = win32com.client.Dispatch("ADODB.Connection")
    conn.CommandTimeout = 0
    conn.Open(f"Provider=MSOLAP;Data Source=localhost:{port};")
    try:
        rs = win32com.client.Dispatch("ADODB.Recordset")
        rs.Open(prep_query(dax), conn)
        cols, rows = _grab(rs)
        if not _has_dim(cols):
            try:
                rs2 = rs.NextRecordset()
                if rs2 is not None:
                    cols, rows = _grab(rs2)
            except Exception:
                pass
        rs.Close()
        return port, cols, rows
    finally:
        if conn.State == 1:
            conn.Close()

# ============================================================
# 3. PHAN TICH CAU TRUC TU RECORDSET (THAY VI DOAN DAX)
# ============================================================
def inner_name(col):
    m = re.findall(r"\[([^\[\]]*)\]", col)
    return m[-1] if m else col

def truthy(v):
    if v is None: return False
    if isinstance(v, str): return v.strip().lower() in ("true", "-1", "1")
    return bool(v)

def num_or_none(v):
    if v is None or v == "": return None
    try: return float(v)
    except (TypeError, ValueError): return v

def is_pct(name):
    n = name.strip()
    return n.startswith('%') or n.startswith('Tỷ') or n.startswith('T.lệ')

# ----- Giữ đúng thứ tự sắp xếp như Power BI Desktop -----
def _is_flag(name):
    return bool(re.match(r'^Is.*Total$', name))

def _sortkey(v):
    # None/rỗng là nhỏ nhất; ASC -> rỗng lên đầu, DESC -> rỗng xuống cuối (giống PBI)
    if v is None or v == "":
        return (0, 0, "")
    if isinstance(v, bool):
        return (1, 0, int(v))
    if isinstance(v, (int, float)):
        return (1, 0, float(v))
    return (1, 1, str(v))

def parse_order_by(dax):
    """Đọc mệnh đề ORDER BY của EVALUATE thân bảng (block cuối cùng).
    Trả về list [(tên_cột, desc_bool), ...] theo thứ tự ưu tiên."""
    blocks = re.split(r'\bORDER\s+BY\b', dax, flags=re.IGNORECASE)
    if len(blocks) < 2:
        return []
    keys = []
    for term in blocks[-1].split(','):
        m = re.search(r"\[([^\]]+)\]", term)
        if not m:
            continue
        keys.append((m.group(1), bool(re.search(r'\bDESC\b', term, re.IGNORECASE))))
    return keys

def sort_rows_like_pbi(dax, cols, rows):
    """Sắp xếp lại rows đúng thứ tự ORDER BY của visual (đa khoá, ổn định)."""
    keys = parse_order_by(dax)
    if not keys:
        return rows
    idx = {inner_name(c): i for i, c in enumerate(cols)}
    out = list(rows)
    for name, desc in reversed(keys):          # stable multi-pass: khoá phụ trước
        if name not in idx:
            continue                            # vd [ColumnIndex] không có trong __DS0Core
        i = idx[name]
        if _is_flag(name):
            out.sort(key=lambda r, i=i: 1 if truthy(r[i]) else 0, reverse=desc)
        else:
            out.sort(key=lambda r, i=i: _sortkey(r[i]), reverse=desc)
    return out

def suppress_repeats(out, ndim):
    """Bỏ lặp nhãn cha (CN, BRN_NAME...) ở các dòng con — giống layout PBI:
    chỉ in nhãn ở dòng đầu của mỗi nhóm, các dòng sau để trống."""
    new, prev = [], None
    for (L, cells, rk, src) in out:
        disp = list(cells)
        if prev is not None:
            cut = 0
            while cut < ndim and disp[cut] == prev[cut] and disp[cut] not in (None, ""):
                cut += 1
            for k in range(cut):
                disp[k] = ""
        new.append((L, disp, rk, src))
        prev = list(cells)
    return new

def parse_structure(dax, cols):
    all_cols = list(cols)
    
    all_dims = []
    for c in all_cols:
        c_inner = inner_name(c)
        # Các cột dimension trả về từ Engine thường chứa ký tự '[' nhưng không bắt đầu bằng '['
        if '[' in c and not c.lstrip().startswith('['):
            all_dims.append(c_inner)
            
    all_flags = [inner_name(c) for c in all_cols if re.match(r'^Is.*Total$', inner_name(c))]
    
    col_dims, col_flags = [], []
    sec_match = re.search(r'__DS0Secondary(?:Base)?\s*=\s*SUMMARIZE\([^,]+,(.*?)\)', dax, re.IGNORECASE | re.DOTALL)
    if sec_match:
        sec_str = sec_match.group(1)
        col_dims = [inner_name(c) for c in re.findall(r"'[^']+'\[([^\]]+)\]", sec_str)]
        col_flags = [re.sub(r'[\"\[\]]', '', f) for f in re.findall(r'\"Is[^"]+Total\"|\[Is[^\]]+Total\]', sec_str)]
        
    row_dims = [d for d in all_dims if d not in col_dims]
    row_flags = [f for f in all_flags if f not in col_flags]
    
    meas_map = {}
    for m in re.finditer(r'"([^"]+)"\s*,\s*(?:\'[^\']+\')?\[([^\]]+)\]', dax):
        meas_map[m.group(1)] = m.group(2)
        
    meas = []
    for c in all_cols:
        nm = inner_name(c)
        if nm not in all_dims and nm not in all_flags and nm != 'ColumnIndex' and not nm.startswith('SortBy_'):
            meas.append((nm, meas_map.get(nm, nm)))
            
    return row_dims, row_flags, col_dims, col_flags, meas

# ============================================================
# 4. SAP XEP HANG (DUNG TRAILING-POP & GIU NGUYEN NATIVE SORT)
# ============================================================
def ordered_rows(total_rows, idx, dims, flags, show_grand=True):
    """Dựng cây hàng phân cấp dựa trên CỜ subtotal (IsXxxTotal), không dựa vào
    việc giá trị dim có rỗng hay không. Trả về (out, sig) — sig là hàm tạo khoá
    nhận dạng duy nhất cho mỗi dòng (dùng để map ô trong pivot).

    Hai lỗi được sửa so với bản cũ:
      1) Khi có dòng Tổng (grand total) thì mọi nhánh con đều nhận cha là (),
         nhưng () không bao giờ được duyệt -> chỉ in ra mỗi dòng Total.
      2) Dòng CHI TIẾT có giá trị dim rỗng (vd 15_CN trống) bị nhầm thành dòng
         subtotal do thuật toán cắt-đuôi -> bị gộp/biến mất. Nay phân biệt
         subtotal vs chi tiết bằng cờ, nên dòng chi tiết rỗng vẫn được giữ.
    """
    ndim = len(dims)
    flag_idx = [idx[f] for f in flags if f in idx]

    def is_total_row(r):
        return any(truthy(r[fi]) for fi in flag_idx)

    def sig(r):
        # Nhận dạng duy nhất 1 dòng = (giá trị các dim) + (chữ ký cờ subtotal)
        return (tuple(r[idx[d]] for d in dims),
                tuple(bool(truthy(r[fi])) for fi in flag_idx))

    def classify(r):
        vals = [r[idx[d]] for d in dims]
        if is_total_row(r):
            level = ndim
            for i, v in enumerate(vals):
                if v in (None, ""):
                    level = i
                    break
            return level, tuple(vals[:level]), True   # subtotal: khoá = tiền tố
        return ndim, tuple(vals), False                # chi tiết: khoá = đường dẫn đầy đủ

    nodes, order = {}, []
    for r in total_rows:
        level, key, it = classify(r)
        if key not in nodes:
            nodes[key] = (level, r, it)
            order.append(key)

    total_keys = {k for k, (lvl, r, it) in nodes.items() if it}

    children = {k: [] for k in nodes}
    roots = []
    for k in order:
        if k == ():
            continue
        parent = None
        for i in range(len(k) - 1, -1, -1):
            if k[:i] in total_keys:
                parent = k[:i]
                break
        if parent is not None:
            children[parent].append(k)
        else:
            roots.append(k)

    out = []
    if show_grand and () in nodes:
        cells = [""] * ndim
        if ndim: cells[0] = "Total"
        out.append((0, cells, sig(nodes[()][1]), nodes[()][1]))

    # Khong sort lai, su dung Native Sort cua recordset
    def walk(k):
        lvl, r, it = nodes[k]
        cells = [""] * ndim
        for i, v in enumerate(k):
            cells[i] = v
        if it and lvl < ndim:
            cells[lvl] = "Total"
        out.append((lvl, cells, sig(r), r))
        for c in children.get(k, []):
            walk(c)

    # Nếu có dòng Tổng thì bắt đầu từ các con của (); nếu không thì từ roots.
    for k in (children[()] if () in nodes else roots):
        walk(k)

    return out, sig

def col_label(t):
    vals = [str(v) for v in t if v not in (None, "")]
    return " / ".join(vals) if vals else "Total"

def _save(wb):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ============================================================
# 5A. STEPPED: hang phan cap + measure lam cot
# ============================================================
def build_stepped(cols, rows, dims, flags, meas, show_grand=True, repeat_header=False):
    idx = {inner_name(c): i for i, c in enumerate(cols)}
    if not dims:
        dims = [inner_name(c) for c in cols if '[' in c and not c.lstrip().startswith('[')]
    dims = [d for d in dims if d in idx]
    meas = [(a, p) for a, p in meas if a in idx]
    out, _ = ordered_rows(rows, idx, dims, flags, show_grand)
    if not repeat_header:
        out = suppress_repeats(out, len(dims))
    return idx, dims, meas, out

def to_excel_stepped(idx, dims, meas, out):
    wb = Workbook(); ws = wb.active; ws.title = "Matrix"
    ndim = len(dims)
    for j, h in enumerate(list(dims) + [p for _, p in meas], 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10); c.fill = HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    ri = 2
    for L, cells, rk, src in out:
        bold = L < ndim
        fill = PatternFill("solid", fgColor=LEVEL_FILL[min(L, len(LEVEL_FILL) - 1)]) if bold else None
        for j, val in enumerate(cells, 1):
            cell = ws.cell(ri, j, val)
            cell.font = Font(bold=bold, size=10); cell.border = BORDER
            cell.alignment = Alignment(horizontal="left")
            if fill: cell.fill = fill
        for k, (alias, pretty) in enumerate(meas):
            val = num_or_none(src[idx[alias]]) if src is not None else None
            cell = ws.cell(ri, ndim + 1 + k)
            if isinstance(val, float):
                cell.value = val
                cell.number_format = "0.00%" if is_pct(pretty) else "#,##0.00"
            elif val is not None:
                cell.value = val
            cell.font = Font(bold=bold, size=10); cell.alignment = Alignment(horizontal="right"); cell.border = BORDER
            if fill: cell.fill = fill
        ri += 1
    for j in range(1, ndim + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16 if j < ndim else 38
    for k in range(len(meas)):
        ws.column_dimensions[get_column_letter(ndim + 1 + k)].width = 15
    ws.freeze_panes = ws.cell(2, ndim + 1).coordinate
    return _save(wb)

def preview_stepped(idx, dims, meas, out):
    records = []
    for L, cells, rk, src in out:
        rec = {d: cells[i] for i, d in enumerate(dims)}
        for alias, pretty in meas:
            v = num_or_none(src[idx[alias]]) if src is not None else None
            rec[pretty] = v if isinstance(v, float) else (v if v is not None else "")
        records.append(rec)
    return pd.DataFrame(records, columns=list(dims) + [p for _, p in meas])

# ============================================================
# 5B. PIVOT: hang phan cap + cot phan cap (Date) x measure
# ============================================================
def build_pivot(cols, rows, row_dims, row_flags, col_dims, meas,
                show_row_total=True, show_col_total=True, repeat_header=False):
    idx = {inner_name(c): i for i, c in enumerate(cols)}
    row_dims = [d for d in row_dims if d in idx]
    col_dims = [d for d in col_dims if d in idx]
    meas = [(a, p) for a, p in meas if a in idx]

    total_rows = [r for r in rows if all(r[idx[c]] in (None, "") for c in col_dims)]
    out, sig = ordered_rows(total_rows, idx, row_dims, row_flags, show_row_total)
    if not repeat_header:
        out = suppress_repeats(out, len(row_dims))

    colkeys_raw = list(dict.fromkeys(tuple(r[idx[c]] for c in col_dims) for r in rows))
    colkeys = [k for k in colkeys_raw if not all(v in (None, "") for v in k)]
    colkeys.sort(key=lambda k: tuple(_sortkey(v) for v in k))   # cột (Date) tăng dần như PBI
    totals = [k for k in colkeys_raw if all(v in (None, "") for v in k)]
    if show_col_total and len(colkeys) > 1: colkeys.extend(totals)   # 1 cột thì tổng cột trùng -> bỏ

    cellmap = {}
    for r in rows:
        ckey = tuple(r[idx[c]] for c in col_dims)
        cellmap[(sig(r), ckey)] = r
    return idx, row_dims, col_dims, meas, out, colkeys, cellmap

def to_excel_pivot(idx, row_dims, col_dims, meas, out, colkeys, cellmap):
    wb = Workbook(); ws = wb.active; ws.title = "Matrix"
    ndim = len(row_dims); nmeas = len(meas)
    two_row = nmeas > 1
    start = (2 if two_row else 1) + 1

    for j, d in enumerate(row_dims, 1):
        c = ws.cell(1, j, d)
        if two_row:
            ws.merge_cells(start_row=1, start_column=j, end_row=2, end_column=j)
        c.font = Font(bold=True, color="FFFFFF", size=10); c.fill = HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER

    j = ndim + 1
    for ck in colkeys:
        label = col_label(ck)
        if two_row:
            ws.merge_cells(start_row=1, start_column=j, end_row=1, end_column=j + nmeas - 1)
            top = ws.cell(1, j, label)
            top.font = Font(bold=True, color="FFFFFF", size=10); top.fill = HDR
            top.alignment = Alignment(horizontal="center", vertical="center"); top.border = BORDER
            for k, (_, pretty) in enumerate(meas):
                cc = ws.cell(2, j + k, pretty)
                cc.font = Font(bold=True, color="FFFFFF", size=9); cc.fill = HDR
                cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cc.border = BORDER
        else:
            cc = ws.cell(1, j, label)
            cc.font = Font(bold=True, color="FFFFFF", size=10); cc.fill = HDR
            cc.alignment = Alignment(horizontal="center", vertical="center"); cc.border = BORDER
        j += nmeas

    ri = start
    for L, cells, rk, src in out:
        bold = L < ndim
        fill = PatternFill("solid", fgColor=LEVEL_FILL[min(L, len(LEVEL_FILL) - 1)]) if bold else None
        for cj, val in enumerate(cells, 1):
            cell = ws.cell(ri, cj, val)
            cell.font = Font(bold=bold, size=10); cell.border = BORDER; cell.alignment = Alignment(horizontal="left")
            if fill: cell.fill = fill
        cj = ndim + 1
        for ck in colkeys:
            srccell = cellmap.get((rk, ck))
            for (alias, pretty) in meas:
                cell = ws.cell(ri, cj)
                val = num_or_none(srccell[idx[alias]]) if srccell is not None else None
                if isinstance(val, float):
                    cell.value = val
                    cell.number_format = "0.00%" if is_pct(pretty) else "#,##0.00"
                elif val is not None:
                    cell.value = val
                cell.font = Font(bold=bold, size=10); cell.alignment = Alignment(horizontal="right"); cell.border = BORDER
                if fill: cell.fill = fill
                cj += 1
        ri += 1

    for jj in range(1, ndim + 1):
        ws.column_dimensions[get_column_letter(jj)].width = 16 if jj < ndim else 36
    for jj in range(ndim + 1, ndim + 1 + len(colkeys) * nmeas):
        ws.column_dimensions[get_column_letter(jj)].width = 14
    ws.freeze_panes = ws.cell(start, ndim + 1).coordinate
    return _save(wb)

def preview_pivot(idx, row_dims, col_dims, meas, out, colkeys, cellmap):
    nmeas = len(meas)
    flat_cols = []
    for ck in colkeys:
        for (_, pretty) in meas:
            flat_cols.append(f"{col_label(ck)} · {pretty}" if nmeas > 1 else col_label(ck))
    columns = list(row_dims) + flat_cols
    records = []
    for L, cells, rk, src in out:
        rec = {d: cells[i] for i, d in enumerate(row_dims)}
        ci = 0
        for ck in colkeys:
            srccell = cellmap.get((rk, ck))
            for (alias, pretty) in meas:
                v = num_or_none(srccell[idx[alias]]) if srccell is not None else None
                rec[flat_cols[ci]] = v if isinstance(v, float) else (v if v is not None else "")
                ci += 1
        records.append(rec)
    return pd.DataFrame(records, columns=columns)

# ============================================================
# 5C. DAO (chi tieu lam hang)
# ============================================================
def build_transpose(cols, rows, dims, flags, meas, show_total=True):
    idx = {inner_name(c): i for i, c in enumerate(cols)}
    if not dims:
        dims = [inner_name(c) for c in cols if '[' in c and not c.lstrip().startswith('[')]
    dim = dims[0]; flag = flags[0] if flags else None
    meas = [(a, p) for a, p in meas if a in idx]
    headers, col_rows = [], []
    for r in rows:
        is_total = truthy(r[idx[flag]]) if (flag and flag in idx) else (r[idx[dim]] in (None, ""))
        if is_total and not show_total:
            continue
        headers.append("Total" if is_total else r[idx[dim]]); col_rows.append(r)
    return idx, dim, meas, headers, col_rows

def to_excel_transpose(idx, dim, meas, headers, col_rows):
    wb = Workbook(); ws = wb.active; ws.title = "Matrix"
    c = ws.cell(1, 1, "Chỉ tiêu")
    c.font = Font(bold=True, color="FFFFFF", size=10); c.fill = HDR; c.border = BORDER
    for j, h in enumerate(headers, 2):
        cc = ws.cell(1, j, h)
        cc.font = Font(bold=True, color="FFFFFF", size=10); cc.fill = HDR
        cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
    for i, (alias, pretty) in enumerate(meas, 2):
        nc = ws.cell(i, 1, pretty); nc.font = Font(bold=True, size=10); nc.border = BORDER
        pct = is_pct(pretty)
        for j, r in enumerate(col_rows, 2):
            val = num_or_none(r[idx[alias]]); cell = ws.cell(i, j)
            if isinstance(val, float):
                cell.value = val; cell.number_format = "0.00%" if pct else "#,##0.00"
            elif val is not None:
                cell.value = val
            cell.alignment = Alignment(horizontal="right"); cell.border = BORDER
    ws.column_dimensions["A"].width = 38
    for j in range(2, len(headers) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.freeze_panes = "B2"
    return _save(wb)

def preview_transpose(idx, dim, meas, headers, col_rows):
    data = {p: [num_or_none(r[idx[a]]) if num_or_none(r[idx[a]]) is not None else "" for r in col_rows]
            for a, p in meas}
    df = pd.DataFrame(data, index=headers).T
    df.index.name = "Chỉ tiêu"
    return df

# ============================================================
# 5D. FALLBACK: xuat phang nguyen data
# ============================================================
def flat_dump(cols, rows, meas):
    pretty = {a: p for a, p in meas}
    names, keep = [], []
    for i, c in enumerate(cols):
        nm = inner_name(c)
        if nm.startswith("SortBy") or nm == "ColumnIndex":
            continue
        names.append(pretty.get(nm, nm)); keep.append(i)
    data = [[r[i] for i in keep] for r in rows]
    df = pd.DataFrame(data, columns=names)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Data")
    buf.seek(0)
    return buf.getvalue(), df

# ============================================================
# 6. DIEU PHOI
# ============================================================
def process(dax, orient, repeat_header=False):
    port, cols, rows = run_query(dax)
    rows = sort_rows_like_pbi(dax, cols, rows)          # giữ đúng thứ tự như PBI Desktop
    row_dims, row_flags, col_dims, col_flags, meas = parse_structure(dax, cols)

    show_row_total = not re.search(r'IsGrandTotalRowTotal\]\s*=\s*FALSE', dax, re.IGNORECASE)
    show_col_total = not re.search(r'IsGrandTotalColumnTotal\]\s*=\s*FALSE', dax, re.IGNORECASE)

    idx0 = {inner_name(c): i for i, c in enumerate(cols)}

    if col_dims and any(d in idx0 for d in col_dims):                       # PIVOT
        idx, rd, cd, m, out, colkeys, cellmap = build_pivot(
            cols, rows, row_dims, row_flags, col_dims, meas,
            show_row_total, show_col_total, repeat_header)
        xlsx = to_excel_pivot(idx, rd, cd, m, out, colkeys, cellmap)
        prev = preview_pivot(idx, rd, cd, m, out, colkeys, cellmap)
        return xlsx, prev, f"Port {port} · pivot · {len(out)} hàng × {len(colkeys)} cột × {len(m)} chỉ tiêu"

    if orient.startswith("Chỉ tiêu"):                                       # DAO
        idx, dim, m, headers, col_rows = build_transpose(
            cols, rows, row_dims, row_flags, meas, show_row_total, repeat_header)
        xlsx = to_excel_transpose(idx, dim, m, headers, col_rows)
        prev = preview_transpose(idx, dim, m, headers, col_rows)
        return xlsx, prev, f"Port {port} · đảo · {len(m)} chỉ tiêu × {len(headers)} cột"

    idx, dims, m, out = build_stepped(                                      # STEPPED
        cols, rows, row_dims, row_flags, meas, show_row_total)
    xlsx = to_excel_stepped(idx, dims, m, out)
    prev = preview_stepped(idx, dims, m, out)
    return xlsx, prev, f"Port {port} · {len(out)} dòng · {len(dims)} cấp · {len(m)} chỉ tiêu"

# ============================================================
# GIAO DIEN
# ============================================================
st.title("📊 PBI Matrix Exporter - Phantichdulieu.vn")
st.caption("Paste query từ **Performance Analyzer** của Power BI Desktop đang chạy → Export → tải Excel đúng định dạng.")

with st.expander("Cách lấy query"):
    st.markdown(
        "1. Power BI Desktop: **Optimize → Performance Analyzer → Start recording → Refresh visuals**, bấm vào bảng.\n"
        "2. Bung dòng visual → **Copy query**.\n"
        "3. Dán vào ô dưới, bấm **Tạo & Export**. (Giữ file .pbix đang mở.)"
    )

dax = st.text_area("Dán DAX query", height=260, placeholder="// DAX Query\nDEFINE ...")

orient = st.radio(
    "Hướng bảng (chỉ áp dụng khi query 1 chiều, không có chiều trên cột)",
    ["Chiều làm hàng (mặc định)", "Chỉ tiêu làm hàng – chiều làm cột (đảo)"],
    horizontal=True,
)

repeat_header = st.checkbox(
    "Lặp lại tiêu đề cha ở mọi dòng",
    value=False,
    help="Bật: lặp lại nhãn cha trên mọi dòng. Tắt: giống Power BI."
)

c1, c2 = st.columns([1, 3])
with c1:
    go = st.button("🚀 Tạo & Export", type="primary", use_container_width=True)
with c2:
    fname = st.text_input("Tên file", value="pbi_matrix.xlsx", label_visibility="collapsed")

if go:
    if not dax.strip():
        st.warning("Bạn chưa dán query.")
    else:
        try:
            with st.spinner("Đang kết nối Power BI và xử lý..."):
                import pythoncom
                pythoncom.CoInitialize()
                try:
                    xlsx, prev, info = process(dax.strip(), orient, repeat_header)
                finally:
                    pythoncom.CoUninitialize()
            st.session_state.update(xlsx=xlsx, preview=prev, fname=fname or "pbi_matrix.xlsx")
            st.success("Xong! " + info)
        except Exception as e:
            try:
                import pythoncom
                pythoncom.CoInitialize()
                try:
                    port, cols, rows = run_query(dax.strip())
                finally:
                    pythoncom.CoUninitialize()
                # Fallback parser
                def fallback_parse(dx, cl):
                    am = [inner_name(c) for c in cl if '[' in c and not c.lstrip().startswith('[')]
                    fl = [inner_name(c) for c in cl if re.match(r'^Is.*Total$', inner_name(c))]
                    mm = {}
                    for m in re.finditer(r'"([^"]+)"\s*,\s*(?:\'[^\']+\')?\[([^\]]+)\]', dx):
                        mm[m.group(1)] = m.group(2)
                    ms = [(inner_name(c), mm.get(inner_name(c), inner_name(c))) for c in cl if inner_name(c) not in am and inner_name(c) not in fl and inner_name(c) != 'ColumnIndex']
                    return am, fl, [], [], ms
                _, _, _, _, meas = fallback_parse(dax, cols)
                xlsx, prev = flat_dump(cols, rows, meas)
                st.session_state.update(xlsx=xlsx, preview=prev, fname=fname or "pbi_matrix.xlsx")
                st.warning(f"Không dựng được matrix ({e}). Đã xuất phẳng nguyên dữ liệu để dùng tạm.")
            except Exception as e2:
                st.error(f"Lỗi: {e2}")

if "xlsx" in st.session_state:
    st.download_button(
        "⬇️ Tải file Excel", data=st.session_state["xlsx"],
        file_name=st.session_state.get("fname", "pbi_matrix.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.dataframe(st.session_state["preview"], use_container_width=True, height=520)